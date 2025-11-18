# header_multirow.py — robust multirow header detection (row+band ML scoring), depth≤5
from __future__ import annotations
import re, math, pickle
from pathlib import Path
from typing import List, Tuple, Optional
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# 탐색 파라미터
HEADER_SCAN_ROWS = 100   # 상단 몇 행까지 헤더 후보로 볼지
MAX_HEADER_DEPTH = 5     # 최대 다단 헤더 깊이

BASE_DIR = Path(__file__).resolve().parent
_BAND_MODEL_PATH = BASE_DIR / "header_band_model.pkl"
_BAND_MODEL: Optional[tuple[str, object]] = None
_BAND_MODEL_FAILED = False

_token_cleaner = re.compile(r"\s+|\n|\r")
_nonword = re.compile(r"[^\w가-힣%\-/.,()]+", re.UNICODE)

def _norm(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = _token_cleaner.sub(" ", str(x)).strip()
    s = _nonword.sub(" ", s)
    return re.sub(r"\s+", " ", s)

def load_sheet_merge_aware(file_path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    ws: Worksheet = wb[sheet_name]
    max_row, max_col = ws.max_row, ws.max_column
    grid = [[None for _ in range(max_col)] for _ in range(max_row)]
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            grid[r-1][c-1] = ws.cell(row=r, column=c).value
    # 병합 복원(값 전파)
    merged_ranges = []
    for mr in ws.merged_cells.ranges:
        v = grid[mr.min_row-1][mr.min_col-1]
        for rr in range(mr.min_row-1, mr.max_row):
            for cc in range(mr.min_col-1, mr.max_col):
                grid[rr][cc] = v
        merged_ranges.append((mr.min_row-1, mr.max_row-1, mr.min_col-1, mr.max_col-1))
    df = pd.DataFrame(grid)
    df.dropna(axis=1, how='all', inplace=True)
    df.dropna(axis=0, how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)
    try:
        df.__dict__["__merges__"] = merged_ranges
    except Exception:
        setattr(df, "__merges__", merged_ranges)
    return df

# ----------------- token / row helpers -----------------
def _token_class(x: object) -> str:
    s = "" if x is None or (isinstance(x, float) and pd.isna(x)) else str(x).strip()
    if s == "": return "EMPTY"
    if re.fullmatch(r"[-+]?\d+(\.\d+)?", s): return "NUM"
    if re.fullmatch(r"\d{4}[-/.]?(\d{1,2})([-/.]?(\d{1,2}))?", s): return "DATE"
    return "TEXT"

def _row_vals(df: pd.DataFrame, r: int):
    return df.iloc[r, :].tolist()

def _nonempty(vals):
    return [v for v in vals if not (isinstance(v, float) and pd.isna(v)) and str(v).strip()!=""]

def _is_data_like_row(df: pd.DataFrame, r: int) -> bool:
    vals = _nonempty(_row_vals(df, r))
    if not vals: return False
    num_ratio = sum(1 for v in vals if _token_class(v) in ("NUM","DATE"))/len(vals)
    return (len(vals) >= max(2, df.shape[1]//4)) and (num_ratio >= 0.30)

def _flip_ratio(df: pd.DataFrame, r: int, d: int) -> float:
    flips = 0; total = 0
    width = min(df.shape[1], 400)  # 초광폭 시트 방지
    for c in range(width):
        h = _token_class(df.iat[r, c])
        dv = _token_class(df.iat[d, c])
        if h == "EMPTY" and dv == "EMPTY": continue
        total += 1
        if h != dv: flips += 1
    return 0.0 if total == 0 else flips/total

def _row_text_num_date(df: pd.DataFrame, r: int):
    vals = _nonempty(_row_vals(df, r))
    if not vals: return 0.0, 0.0, 0.0, 0.0, 0.0
    t = n = d = 0
    lens = []
    for v in vals:
        cls = _token_class(v)
        if cls == "TEXT": t += 1
        elif cls == "NUM": n += 1
        elif cls == "DATE": d += 1
        s = _norm(v)
        if s: lens.append(len(s))
    total = max(1, t+n+d)
    uniq = len(set(_norm(v).lower() for v in vals if _norm(v)!="")) / max(1, len(vals))
    empt = 1 - (len(vals) / max(1, df.shape[1]))
    avg_len = np.mean(lens) if lens else 0.0
    return t/total, n/total, d/total, uniq, min(max(empt, 0.0), 1.0), avg_len

def _row_merge_bonus(df: pd.DataFrame, r: int) -> float:
    merges = getattr(df, "__merges__", []) or []
    bonus = 0.0
    for (rmin, rmax, cmin, cmax) in merges:
        if rmin <= r <= rmax:  # 이 행이 병합 범위에 걸려 있다
            h = (rmax - rmin + 1); w = (cmax - cmin + 1)
            if w >= 2: bonus += min(0.5, 0.05*(w-1))
            if h >= 2: bonus += min(0.5, 0.08*(h-1))
    return min(1.0, bonus)

def detect_data_start_strict(df: pd.DataFrame) -> int:
    """
    데이터로 보이는 첫 행을 찾는다.
    두 줄 이상 연속한 데이터 패턴을 우선하며, 실패 시 최초 데이터형 행을 반환한다.
    """
    if df is None or len(df) == 0:
        return 0
    scan = min(len(df), HEADER_SCAN_ROWS)
    streak_start: Optional[int] = None
    for r in range(scan):
        if _is_data_like_row(df, r):
            if streak_start is None:
                streak_start = r
            elif r - streak_start >= 1:
                return streak_start
        else:
            streak_start = None
    for r in range(scan):
        if _is_data_like_row(df, r):
            return r
    return min(scan, len(df) - 1)

# ----------------- row-level logistic -----------------
# 행 단위 헤더성 점수(사전 튜닝 계수). 텍스트↑, 숫자/날짜↓, 유니크↑, 공란↓, 평균길이 적당, 병합 보너스, 아래행 flip 보너스
_ROW_COEF = np.array([[ 2.2, -0.9, -0.6, 0.8, -1.0, 0.3, 0.7, 0.9 ]], dtype=float)
_ROW_INTER = np.array([-0.7], dtype=float)

def _sigmoid(z): return 1.0/(1.0 + np.exp(-z))

def _row_score(df: pd.DataFrame, r: int) -> float:
    tr, nr, dr, uniq, empt, avg_len = _row_text_num_date(df, r)
    # 평균 길이 6~18자 선호(너무 짧거나 길면 감점)
    len_center = math.exp(-abs(avg_len - 10.0)/12.0)
    # 아래 첫 데이터행이 존재하면 flip 보너스
    d = None
    for rr in range(r+1, min(len(df), r+40)):
        if _nonempty(_row_vals(df, rr)) and _is_data_like_row(df, rr):
            d = rr; break
    flip = _flip_ratio(df, r, d) if d is not None else 0.0
    mergeb = _row_merge_bonus(df, r)
    x = np.array([tr, nr, dr, uniq, empt, len_center, flip, mergeb]).reshape(1,-1)
    z = float(x @ _ROW_COEF.T + _ROW_INTER)
    return _sigmoid(z)

# ----------------- band-level scoring -----------------
def _band_internal_diversity(df: pd.DataFrame, start: int, end: int) -> float:
    """밴드 내에서 컬럼별로 '윗줄/아랫줄이 모두 비어있지 않고 서로 다름' 비율 → 다단 구조일수록↑"""
    if end <= start: return 0.0
    rows = [ [_norm(v) for v in _row_vals(df, r)] for r in range(start, end+1) ]
    rows = [r for r in rows if any(cell != "" for cell in r)]
    if len(rows) < 2: return 0.0
    cols = min(len(rows[0]), 600)
    diff = 0; base = 0
    for c in range(cols):
        col_vals = [rows[k][c] for k in range(len(rows))]
        nz = [v for v in col_vals if v != ""]
        if len(nz) >= 2:
            base += 1
            if len(set(nz)) >= 2:
                diff += 1
    return diff / max(1, base)

def _merge_presence_bonus(df: pd.DataFrame, r0: int, r1: int) -> float:
    merges = getattr(df, "__merges__", []) or []
    bonus = 0.0
    for (rmin, rmax, cmin, cmax) in merges:
        if not (rmax < r0 or rmin > r1):
            h = (rmax - rmin + 1); w = (cmax - cmin + 1)
            if w >= 2: bonus += min(0.6, 0.06*(w-1))
            if h >= 2: bonus += min(0.6, 0.10*(h-1))
    return min(1.2, bonus)

def _band_core_features(df: pd.DataFrame, start: int, end: int) -> np.ndarray:
    # 아래 첫 데이터 행
    d = None
    for rr in range(end+1, min(len(df), end+60)):
        if _nonempty(_row_vals(df, rr)) and _is_data_like_row(df, rr):
            d = rr; break
    # 행 점수 평균
    row_scores = [_row_score(df, r) for r in range(start, end+1)]
    avg_row = float(np.mean(row_scores))
    min_row = float(np.min(row_scores))
    depth = (end - start + 1)
    depth_norm = depth / MAX_HEADER_DEPTH
    flip = _flip_ratio(df, end, d) if d is not None else 0.0
    diversity = _band_internal_diversity(df, start, end)  # 다단이면↑
    merges = _merge_presence_bonus(df, start, end)        # 병합 흔적이면↑
    # 밴드 자체의 공란/유니크 지표
    cells = []
    for rr in range(start, end+1):
        cells.extend(_row_vals(df, rr))
    toks = [_norm(v) for v in cells]
    toks_nz = [t for t in toks if t != ""]
    uniq = len(set(toks_nz))/max(1, len(toks_nz))
    empt = 1 - (len(toks_nz)/max(1, len(toks)))
    return np.array([avg_row, min_row, depth_norm, flip, diversity, merges, uniq, empt], dtype=float)

def _data_follow_ratio(df: pd.DataFrame, end: int, window: int = 5) -> float:
    total = 0
    hits = 0
    for rr in range(end+1, min(len(df), end+1+window)):
        total += 1
        if _is_data_like_row(df, rr):
            hits += 1
    return hits / max(1, total)

def band_features(df: pd.DataFrame, start: int, end: int, data_start: Optional[int] = None) -> np.ndarray:
    core = _band_core_features(df, start, end)
    ds = data_start if data_start is not None else detect_data_start_strict(df)
    if ds is None:
        ds = 0
    ds = max(0, min(ds, len(df)))  # clamp
    base = max(1, ds + 1)
    start_norm = max(0.0, min(1.0, start / base))
    end_norm = max(0.0, min(1.0, end / base))
    gap = max(0.0, (ds - end - 1) / base)
    follow = _data_follow_ratio(df, end)
    extras = np.array([start_norm, end_norm, gap, follow], dtype=float)
    return np.concatenate([core, extras])

# 사전 튜닝 계수(행 평균↑, 최소행도 너무 낮지 않게, 깊이↑, flip↑, 다양성↑, 병합↑, 유니크↑, 공란↓)
_BAND_COEF = np.array([[ 2.6, 0.6, 1.2, 1.4, 1.3, 1.0, 0.4, -0.9 ]], dtype=float)
_BAND_INTER = np.array([-1.1], dtype=float)

def _load_band_model() -> Optional[tuple[str, object]]:
    global _BAND_MODEL, _BAND_MODEL_FAILED
    if _BAND_MODEL is not None:
        return _BAND_MODEL
    if _BAND_MODEL_FAILED:
        return None
    try:
        if not _BAND_MODEL_PATH.exists():
            _BAND_MODEL_FAILED = True
            return None
        with open(_BAND_MODEL_PATH, "rb") as f:
            impl = pickle.load(f)
        if isinstance(impl, tuple) and len(impl) == 2:
            _BAND_MODEL = (impl[0], impl[1])
            return _BAND_MODEL
    except Exception:
        pass
    _BAND_MODEL_FAILED = True
    return None

def _band_score(df: pd.DataFrame, start: int, end: int, data_start: Optional[int] = None) -> float:
    feats = band_features(df, start, end, data_start=data_start)
    impl = _load_band_model()
    if impl is not None:
        model_type, model = impl
        x = np.array(feats, dtype=float).reshape(1, -1)
        try:
            if model_type == "lightgbm":
                proba = model.predict(x)
                score = float(np.ravel(proba)[0])
                return min(max(score, 0.0), 1.0)
            elif model_type == "sk_hgb":
                proba = model.predict_proba(x)
                score = float(proba[0, 1])
                return min(max(score, 0.0), 1.0)
        except Exception:
            pass
    base = feats[:_BAND_COEF.shape[1]].reshape(1, -1)
    z = float(base @ _BAND_COEF.T + _BAND_INTER)
    return _sigmoid(z)

def detect_header_band_and_build(df: pd.DataFrame) -> Tuple[int,int,List[str]]:
    """
    헤더 밴드(연속 1~5행)를 찾아 (start, end, headers)를 반환.
    headers는 위→아래를 ' | '로 합친 다단 컬럼명.
    """
    scan = min(HEADER_SCAN_ROWS, len(df))
    data_start = detect_data_start_strict(df)
    search_limit = min(scan, max(data_start + 2, MAX_HEADER_DEPTH + 1, 10))
    best = None
    best_score = -1e9

    # 1) 모든 밴드 후보 스코어링
    candidates = []
    for r in range(search_limit):
        for k in range(1, min(MAX_HEADER_DEPTH, search_limit - r) + 1):
            end = r + k - 1
            sc = _band_score(df, r, end, data_start=data_start)
            # 너무 아래에 있으면 약한 패널티
            sc -= 0.02 * max(0, r - 6)
            if end >= data_start:
                sc -= 0.05 * (end - data_start + 1)
            candidates.append((sc, r, end))
            if sc > best_score:
                best = (r, end); best_score = sc

    # 2) 1행 vs 2행 밴드 경합 시, 2행이 충분히 근접하면 2행을 우선(다단 보정)
    if best:
        r0, e0 = best
        k0 = e0 - r0 + 1
        if k0 == 1:
            # 같은 시작 r0에 대해 k=2가 존재하면 비교
            alt = [(sc, r, e) for (sc, r, e) in candidates if r == r0 and (e - r + 1) == 2]
            if alt:
                sc2, r2, e2 = max(alt, key=lambda x: x[0])
                if sc2 >= best_score * 0.97:  # 2행 밴드가 97% 이상 근접하면 다단으로 선택
                    best = (r2, e2)
                    best_score = sc2

    # 3) 비슷한 점수 내에 더 위쪽 후보가 있으면 우선
    if best and candidates:
        tolerance = 0.035
        target_start = best[0]
        near = [
            c for c in candidates
            if c[0] >= best_score - tolerance and c[1] <= target_start - 2
        ]
        if near:
            near.sort(key=lambda item: (item[1], -(item[2] - item[1]), -item[0]))
            chosen = near[0]
            best = (chosen[1], chosen[2])
            best_score = chosen[0]

    if best is None:
        # fallback: 1행 헤더
        headers = [(_norm(v) or f"컬럼{c+1}") for c, v in enumerate(df.iloc[0, :].tolist())]
        out=[]; used={}
        for s in headers:
            k = used.get(s,0); used[s]=k+1
            out.append(s if k==0 else f"{s}_{k+1}")
        return (0, 0, out)

    start, end = best

    # 3) 다단 헤더 합성 (빈칸은 건너뛰고 상·하위만 체인)
    bands = [df.iloc[r, :].apply(_norm).tolist() for r in range(start, end+1)]
    combined = []
    for c in range(len(bands[0])):
        parts = [bands[r][c] for r in range(len(bands)) if bands[r][c] != ""]
        name = " | ".join(parts).strip() if parts else ""
        combined.append(name or f"컬럼{c+1}")

    # 4) 중복 방지
    out=[]; used={}
    for s in combined:
        k = used.get(s,0); used[s]=k+1
        out.append(s if k==0 else f"{s}_{k+1}")
    return (start, end, out)

def build_multirow_headers(df: pd.DataFrame, _hint: int) -> List[str]:
    _, _, headers = detect_header_band_and_build(df)
    return headers
