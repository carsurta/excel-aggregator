# sheet_match.py — logistic recommender with header similarity
from __future__ import annotations
import numpy as np
from difflib import SequenceMatcher
from typing import List, Dict
from sklearn.linear_model import LogisticRegression
from openpyxl import load_workbook

# -------- util: normalize --------
def _norm_name(s: str) -> str:
    if not s: return ""
    return "".join(ch for ch in s.lower().strip() if ch.isalnum())

def _norm_cols(cols: List[str]) -> List[str]:
    out = []
    for c in cols or []:
        s = _norm_name(c)
        if s: out.append(s)
    return out

# -------- name features --------
def _name_features(a: str, b: str) -> np.ndarray:
    A, B = _norm_name(a), _norm_name(b)
    if not A or not B:
        return np.zeros(6)
    sm = SequenceMatcher(None, A, B)
    ratio = sm.ratio()
    prefix = 1.0 if A.startswith(B) or B.startswith(A) else 0.0
    suffix = 1.0 if A.endswith(B) or B.endswith(A) else 0.0
    jaccard = len(set(A) & set(B)) / max(1, len(set(A) | set(B)))
    overlap = len(set(A) & set(B)) / max(1, len(set(A)))
    length_diff = abs(len(A) - len(B)) / max(len(A), len(B))
    return np.array([ratio, prefix, suffix, jaccard, overlap, length_diff])

# -------- header features --------
def _header_features(base_cols: List[str], cand_cols: List[str]) -> np.ndarray:
    B = _norm_cols(base_cols)
    C = _norm_cols(cand_cols)
    if not B or not C:
        # [jaccard, coverageB, coverageC, lcs_ratio]
        return np.zeros(4)
    setB, setC = set(B), set(C)
    jacc = len(setB & setC) / max(1, len(setB | setC))
    covB = len(setB & setC) / max(1, len(setB))
    covC = len(setB & setC) / max(1, len(setC))
    # 순서감(대략): B를 기준으로 C에서 공통열들의 인덱스 순서 LCS 비율
    idxB = {v:i for i,v in enumerate(B)}
    seq = [idxB[v] for v in C if v in idxB]
    # LCS 길이 ≈ 가장 긴 증가 부분수열(LIS) 길이
    lcs = 0
    tails = []
    import bisect
    for x in seq:
        k = bisect.bisect_left(tails, x)
        if k == len(tails): tails.append(x)
        else: tails[k] = x
    lcs = len(tails)
    lcs_ratio = lcs / max(1, min(len(B), len(C)))
    return np.array([jacc, covB, covC, lcs_ratio])

# -------- logistic model (pre-tuned priors) --------
class _Recommender:
    def __init__(self):
        self.model = LogisticRegression()
        # [name 6] + [header 4] = 10차원
        # 가중치: 이름보다 헤더 유사도에 더 강한 비중
        self.model.coef_ = np.array([[ 2.0, 0.6, 0.6, 0.8, 0.7, -1.0,   # name
                                       3.0, 2.0, 1.5, 1.6 ]])           # header
        self.model.intercept_ = np.array([-1.2])
        self.model.classes_ = np.array([0,1])

    def score(self, base_name: str, new_name: str, base_cols: List[str], cand_cols: List[str]) -> float:
        xn = _name_features(base_name, new_name)
        xh = _header_features(base_cols, cand_cols)
        x = np.concatenate([xn, xh]).reshape(1, -1)
        return float(self.model.predict_proba(x)[0, 1])

_rec = _Recommender()

# -------- API --------
def list_sheet_names(file_path: str) -> List[str]:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    return wb.sheetnames

def auto_match_with_headers(
    base: List[tuple[str, List[str]]],   # [(base_sheet_name, base_headers), ...]
    candidates: List[tuple[str, List[str]]],  # [(cand_sheet_name, cand_headers), ...]
    threshold: float = 0.55
) -> List[str]:
    """
    base: 첫 파일에서 사용자가 선택한 (시트명, 헤더열목록)
    candidates: 새 파일의 (시트명, 헤더열목록) 목록
    return: 추천할 시트명 목록 (여러 개 가능)
    """
    selected: Dict[str, float] = {}
    for bname, bcols in base:
        for cname, ccols in candidates:
            p = _rec.score(bname, cname, bcols, ccols)
            if p >= threshold:
                selected[cname] = max(selected.get(cname, 0.0), p)
    # 확률 높은 순으로 정렬해두되, 반환은 시트명 리스트
    return [name for name, _ in sorted(selected.items(), key=lambda x: -x[1])]
