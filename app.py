# -*- coding: utf-8 -*-
# app.py – Excel Aggregator (Drag & Drop) – FINAL UI/Features
from __future__ import annotations
import sys, os, traceback, importlib
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

def _pd():
    return importlib.import_module("pandas")

from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, QTimer, QItemSelection, QItemSelectionModel
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QListWidget,
    QFileDialog, QHBoxLayout, QMessageBox, QTableView, QProgressBar, QCheckBox,
    QSplitter, QListWidgetItem, QDialog, QDialogButtonBox, QListWidget, QSpinBox
)

from header_multirow import (
    load_sheet_merge_aware,
    detect_header_band_and_build,
    build_headers_from_band,
)
from sheet_match import list_sheet_names, auto_match_with_headers
from validators import compute_violations

# ======= Meta columns (Korean, official tone) =======
META_MAP = {
    "file": "출처 파일명",
    "sheet": "출처 시트명",
    "violations": "유효성 점검 메모",
}
META_COLS_KR = list(META_MAP.values())

class _FileRow(QWidget):
    def __init__(self, parent, path: str, on_remove, on_choose, on_header):
        super().__init__(parent)
        self.path = path
        from PyQt6.QtWidgets import QHBoxLayout, QPushButton, QLabel, QSizePolicy
        from PyQt6.QtCore import Qt

        hl = QHBoxLayout(self)
        hl.setContentsMargins(6, 2, 6, 2)
        hl.setSpacing(6)

        # 좌측 X 버튼
        self.btn_x = QPushButton("✕")
        self.btn_x.setFixedWidth(28)
        self.btn_x.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_x.clicked.connect(lambda: on_remove(self.path))
        hl.addWidget(self.btn_x)

        # 가운데 경로 라벨(길면 말줄임)
        self.lbl = QLabel(path)
        self.lbl.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.lbl.setToolTip(path)
        hl.addWidget(self.lbl, 1)

        # 우측 시트 버튼
        self.btn_sheet = QPushButton("시트 선택")
        self.btn_sheet.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_sheet.clicked.connect(lambda: on_choose(self.path, True))
        hl.addWidget(self.btn_sheet)

        self.btn_header = QPushButton("헤더 확인/조정")
        self.btn_header.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_header.clicked.connect(lambda: on_header(self.path))
        hl.addWidget(self.btn_header)

# ============== Qt Model for pandas ==================
class PandasModel(QAbstractTableModel):
    def __init__(self, df):
        super().__init__()
        self._df = df

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._df)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._df.columns)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            val = self._df.iat[index.row(), index.column()]
            try:
                import pandas as pd
                if pd.isna(val):
                    return ""
            except Exception:
                pass
            return "" if val is None else str(val)
        return None

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole):
        if role != Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == Qt.Orientation.Horizontal:
            try:
                return str(self._df.columns[section])
            except Exception:
                return ""
        else:
            return str(section + 1)

# ================= Data structures ====================
@dataclass
class ParsedSheet:
    file_path: str
    sheet_name: str
    columns: List[str]
    data: "pandas.DataFrame"

# ==================== Main UI =========================
class ExcelAggregator(QWidget):

    def _on_meta_label_clicked(self, event):
        self.chk_meta.toggle()
        self._refresh_preview()
        
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Aggregator – Drag & Drop")
        self.setAcceptDrops(True)
        self.resize(1200, 780)

        # preferences: sheet names chosen in the first file
        self.pref_sheet_names: List[str] = []  # used for auto-preselect on later files
        # selected sheets per file
        self.file_sheets: Dict[str, List[str]] = {}
        # 저장된 헤더 시그니처(사용자가 확정한 헤더 행/데이터)
        self.saved_header_signature = None
        self.saved_header_band = None
        self.sheet_header_cache: Dict[Tuple[str, str], Tuple[Tuple[str, ...], Tuple[int, int], List[str]]] = {}

        # ---- Layout: vertical splitter (resizable list + preview) ----
        root = QVBoxLayout(self)
        self.info = QLabel("xlsx 파일/폴더를 드래그 앤 드롭하세요.")
        self.info.setStyleSheet("font-size:14px;color:#333;margin:6px 4px;")
        root.addWidget(self.info)

        splitter = QSplitter(Qt.Orientation.Vertical)
        root.addWidget(splitter, 1)

        # Top: file list area (with Clear & per-item remove via Del key)
        top_widget = QWidget(); top_layout = QVBoxLayout(top_widget); top_layout.setContentsMargins(0,0,0,0)
        self.list = QListWidget()
        self.list.setStyleSheet("QListWidget{border:1px dashed #999;padding:8px;}")
        top_layout.addWidget(self.list, 1)

        hl = QHBoxLayout(); hl.setContentsMargins(4,4,4,0)
        self.btn_add = QPushButton("파일 불러오기")
        self.btn_add.clicked.connect(self._browse_files)
        hl.addWidget(self.btn_add)
        self.btn_clear = QPushButton("모두 제거")
        self.btn_clear.clicked.connect(self._clear)
        hl.addWidget(self.btn_clear)
        hl.addStretch(1)
        top_layout.addLayout(hl)

        splitter.addWidget(top_widget)

        # Bottom: preview table + bottom bar (meta checkbox + Save)
        bottom_widget = QWidget(); bottom_layout = QVBoxLayout(bottom_widget); bottom_layout.setContentsMargins(0,0,0,0)
        self.progress = QProgressBar(); self.progress.setVisible(False)
        bottom_layout.addWidget(self.progress)

        self.table = QTableView()
        bottom_layout.addWidget(self.table, 1)

        bottom_bar = QHBoxLayout()
        self.chk_meta = QCheckBox(); self.chk_meta.setChecked(False)
        self.lbl_meta = QLabel("원본 파일 정보 표시")
        self.lbl_meta.setStyleSheet("padding-left:4px;")
        self.lbl_meta.setCursor(Qt.CursorShape.PointingHandCursor)
        tooltip = (
            "출력 결과물에 다음 정보를 표기합니다:\n"
            f"• {META_MAP['file']}: 취합의 출처가 된 원본 엑셀 파일 이름 (예: 자료제출_부서A.xlsx)\n"
            f"• {META_MAP['sheet']}: 해당 데이터가 있었던 원본 시트 이름 (예: 제출본, 입력데이터)\n"
            f"• {META_MAP['violations']}: 전화/이메일 형식 오류, 빈값, 중복 등 간단한 유효성 점검 메모 (예: [연락처:전화형식] [이메일:형식] [번호:중복])\n\n"
            "체크를 해제하면 위 정보는 출력 파일에 포함되지 않으며 미리보기에서도 숨겨집니다."
        )
        self.chk_meta.setToolTip(tooltip); self.lbl_meta.setToolTip(tooltip)
        self.lbl_meta.mousePressEvent = self._on_meta_label_clicked
        self.chk_meta.stateChanged.connect(self._refresh_preview)
        bottom_bar.addWidget(self.chk_meta)
        bottom_bar.addWidget(self.lbl_meta)

        self.chk_log = QCheckBox()
        self.chk_log.setChecked(False)
        self.lbl_log = QLabel("로그 저장")
        self.lbl_log.setStyleSheet("padding-left:4px;")
        self.lbl_log.setCursor(Qt.CursorShape.PointingHandCursor)
        self.lbl_log.mousePressEvent = lambda e: self.chk_log.toggle()
        bottom_bar.addWidget(self.chk_log)
        bottom_bar.addWidget(self.lbl_log)

        bottom_bar.addStretch(1)

        self.btn_reheader = QPushButton("헤더 다시 선택")
        self.btn_reheader.clicked.connect(self._reselect_header)
        bottom_bar.addWidget(self.btn_reheader)

        self.btn_save = QPushButton("출력 저장…")
        self.btn_save.clicked.connect(self._save_output)
        bottom_bar.addWidget(self.btn_save)
        bottom_layout.addLayout(bottom_bar)

        splitter.addWidget(bottom_widget)
        splitter.setStretchFactor(0, 0)  # top list minimal height by default
        splitter.setStretchFactor(1, 1)  # table expands

        # Data
        self.file_paths: List[str] = []
        self.parsed: List[ParsedSheet] = []
        self.combined = None  # pandas.DataFrame
        self.pref_sheet_names: List[str] = []        # 첫 파일에서 선택한 시트명들
        self.pref_headers: Dict[str, List[str]] = {} # 첫 파일에서 선택한 시트별 헤더 (합친 열명 리스트)

        # Remove key for selected list items
        self.list.keyPressEvent = self._list_keypress

    # ---------- Drag & Drop ----------
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        added_files = []
        for u in urls:
            p = u.toLocalFile()
            if os.path.isdir(p):
                for root, _, files in os.walk(p):
                    for f in files:
                        if f.lower().endswith(".xlsx") and not f.startswith("~$"):
                            added_files.append(os.path.join(root, f))
            elif p.lower().endswith(".xlsx") and not os.path.basename(p).startswith("~$"):
                added_files.append(p)
        if not added_files:
            return

        # For each file, run sheet chooser with auto-preselection based on pref_sheet_names
        for path in added_files:
            self._open_sheet_chooser_for(path)

    def _list_keypress(self, event):
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            for it in self.list.selectedItems():
                w = self.list.itemWidget(it)
                if w and hasattr(w, "path"):
                    self._remove_file(w.path)
        else:
            try:
                from PyQt6.QtWidgets import QListWidget
                QListWidget.keyPressEvent(self.list, event)
            except Exception:
                pass

    def _browse_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "엑셀 파일 선택",
            "",
            "Excel Files (*.xlsx)"
        )
        for path in paths or []:
            if path and path.lower().endswith(".xlsx") and not os.path.basename(path).startswith("~$"):
                self._open_sheet_chooser_for(path)

    # ---------- Sheet selection ----------
    def _add_with_sheet_selection(self, path: str):
        try:
            sheets = list_sheet_names(path)
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "시트 읽기 오류", f"{os.path.basename(path)}: 시트 목록을 읽을 수 없습니다.\n{e}")
            return

        # auto-preselect: match to preferred names if exists
        preselect = []
        if self.pref_sheet_names:
            base_norm = [s.strip().lower() for s in self.pref_sheet_names]
            for s in sheets:
                sn = s.strip().lower()
                if any(b in sn or sn in b for b in base_norm):
                    preselect.append(s)

        dlg = SheetChooser(self, path, sheets, preselect)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        chosen = dlg.selected_sheets()
        if not chosen:
            return

        if not self.file_paths and not self.pref_sheet_names:
            self.pref_sheet_names = chosen[:]

        if path not in self.file_paths:
            self.file_paths.append(path)
            self.file_sheets[path] = chosen
            self.list.addItem(path)
        else:
            self.file_sheets[path] = chosen

    def _add_list_item(self, path: str):
        from PyQt6.QtWidgets import QListWidgetItem
        # 이미 있으면 무시(혹은 덮어쓰기)
        for i in range(self.list.count()):
            it = self.list.item(i)
            w = self.list.itemWidget(it)
            if w and getattr(w, "path", "") == path:
                return it  # 이미 존재
        it = QListWidgetItem(self.list)
        w = _FileRow(self, path, self._remove_file, self._open_sheet_chooser_for, self._open_header_adjust_for)
        it.setSizeHint(w.sizeHint())
        self.list.addItem(it)
        self.list.setItemWidget(it, w)
        return it

    def _refresh_all_items(self):
        # 목록 전체를 현재 self.file_paths 순서대로 재구성
        self.list.clear()
        for p in self.file_paths:
            self._add_list_item(p)

    def _open_header_adjust_for(self, path: str):
        sheets = self.file_sheets.get(path) or []
        if not sheets:
            try:
                sheets = list_sheet_names(path)
            except Exception as e:
                traceback.print_exc()
                QMessageBox.critical(self, "시트 조회 오류", f"{os.path.basename(path)} 시트 목록을 불러올 수 없습니다.\n{e}")
                return
        sheet = None
        if len(sheets) == 1:
            sheet = sheets[0]
        else:
            dlg = SingleSheetDialog(self, path, sheets)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            sheet = dlg.selected_sheet()
        if not sheet:
            return
        try:
            raw = load_sheet_merge_aware(path, sheet)
            st, ed, headers = detect_header_band_and_build(raw)
            # sheet 캐시가 있으면 초기 밴드/시그니처로 활용
            cached = self.sheet_header_cache.get((path, sheet))
            if cached:
                _, band, _ = cached
                st, ed = band
            st, ed, headers = self._resolve_header_band(raw, path, sheet, st, ed, force_dialog=True)
            # 헤더가 갱신되었으니 전체 미리보기 재계산
            self._parse_and_preview()
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "헤더 보정 오류", f"{os.path.basename(path)} / {sheet}\n{e}")

    def _remove_file(self, path: str):
        self.file_paths = [p for p in self.file_paths if p != path]
        if path in self.file_sheets:
            del self.file_sheets[path]
        self.sheet_header_cache = {k: v for k, v in self.sheet_header_cache.items() if k[0] != path}
        self._refresh_all_items()
        self._parse_and_preview()

    def _has_preview(self) -> bool:
        return bool(self.parsed)

    def _header_signature(self, df, start: int, end: int):
        sig: List[str] = []
        for r in range(start, end + 1):
            try:
                row = df.iloc[r, :].tolist()
            except Exception:
                continue
            parts = []
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if s and s.lower() != "nan":
                    parts.append(s)
            sig.append("|".join(parts))
        return tuple(sig)

    def _header_similarity(self, sig_a: Tuple[str, ...], sig_b: Tuple[str, ...]) -> float:
        if not sig_a or not sig_b:
            return 0.0
        def to_tokens(sig):
            toks = []
            for row in sig:
                for part in str(row).lower().split("|"):
                    part = part.strip()
                    if part:
                        toks.append(part)
            return set(toks)
        ta, tb = to_tokens(sig_a), to_tokens(sig_b)
        if not ta or not tb:
            return 0.0
        inter = len(ta & tb)
        union = len(ta | tb)
        return inter / union if union else 0.0

    def _find_best_band(self, df, target_sig: Tuple[str, ...], top_rows: int = 30, max_height: int = 5):
        best_band = None
        best_sig = ()
        best_score = -1.0
        rows = min(len(df), top_rows)
        target_h = max(1, len(target_sig) if target_sig else 1)
        min_h = max(1, target_h - 1)
        max_h = max(min_h, min(max_height, target_h + 1))
        for s in range(rows):
            for e in range(s + min_h - 1, min(rows, s + max_h)):
                sig = self._header_signature(df, s, e)
                score = self._header_similarity(target_sig, sig)
                if score > best_score:
                    best_score = score
                    best_band = (s, e)
                    best_sig = sig
        return best_band, best_sig, best_score

    def _resolve_header_band(
        self,
        df,
        path: str,
        sheet: str,
        guess_start: int,
        guess_end: int,
        force_dialog: bool = False,
    ) -> tuple[int, int, List[str]]:
        key = (path, sheet)
        saved_sheet = self.sheet_header_cache.get(key)
        saved_sig_global = self.saved_header_signature
        saved_band_global = self.saved_header_band

        # 1) sheet별 캐시 우선 적용
        if not force_dialog and saved_sheet:
            sig_sheet, band_sheet, _ = saved_sheet
            if band_sheet[1] < len(df):
                sig_same = self._header_signature(df, band_sheet[0], band_sheet[1])
                if sig_same == sig_sheet:
                    headers = build_headers_from_band(df, band_sheet[0], band_sheet[1])
                    return band_sheet[0], band_sheet[1], headers

        # 2) 글로벌 시그니처 자동 적용 (첫 파일 기준)
        if not force_dialog and saved_sig_global:
            if saved_band_global and saved_band_global[1] < len(df):
                sig_same_band = self._header_signature(df, saved_band_global[0], saved_band_global[1])
                if sig_same_band == saved_sig_global:
                    headers = build_headers_from_band(df, saved_band_global[0], saved_band_global[1])
                    return saved_band_global[0], saved_band_global[1], headers
            sig_guess = self._header_signature(df, guess_start, guess_end)
            if sig_guess == saved_sig_global:
                self.saved_header_band = (guess_start, guess_end)
                headers = build_headers_from_band(df, guess_start, guess_end)
                return guess_start, guess_end, headers

        # 3) 사전 추정 밴드 결정
        pre_start, pre_end = guess_start, guess_end
        target_sig = None
        if saved_sheet:
            target_sig, band_sheet, _ = saved_sheet
            pre_start, pre_end = band_sheet
        elif saved_sig_global:
            target_sig = saved_sig_global

        if target_sig:
            best_band, _, _ = self._find_best_band(df, target_sig)
            if best_band:
                pre_start, pre_end = best_band

        # 4) 보정 UI
        dlg = HeaderAdjustDialog(self, df, pre_start, pre_end)
        accepted = dlg.exec() == QDialog.DialogCode.Accepted
        if accepted:
            start, end = dlg.selected_band()
        else:
            start, end = pre_start, pre_end

        headers = build_headers_from_band(df, start, end)
        sig_final = self._header_signature(df, start, end)

        # sheet별 캐시 저장
        self.sheet_header_cache[key] = (sig_final, (start, end), headers)
        # 글로벌 기본 시그니처는 최초 또는 사용자가 수동 확정한 경우만 갱신
        if self.saved_header_signature is None or accepted:
            self.saved_header_signature = sig_final
            self.saved_header_band = (start, end)
        return start, end, headers

    def _record_pref_from(self, path: str, sheets: list[str]):
        if self.file_paths or self.pref_sheet_names or not sheets:
            return
        self.pref_sheet_names = sheets[:]
        self.pref_headers = {}
        for s in sheets:
            try:
                raw0 = load_sheet_merge_aware(path, s)
                _st, _ed, _hdrs = detect_header_band_and_build(raw0)
                self.pref_headers[s] = _hdrs
            except Exception:
                self.pref_headers[s] = []

    def _finalize_file_selection(self, path: str, chosen: list[str]):
        self.file_sheets[path] = chosen[:]
        if path not in self.file_paths:
            self.file_paths.append(path)
            self._add_list_item(path)

    def _open_sheet_chooser_for(self, path: str, force_dialog: bool = False):
        try:
            sheets = list_sheet_names(path)
        except Exception as e:
            traceback.print_exc()
            QMessageBox.critical(self, "시트 읽기 오류", f"{os.path.basename(path)}: 시트 목록을 불러올 수 없습니다.\n{e}")
            return

        # 1) 각 시트의 헤더 미리 추출
        cand_name_headers = []
        for s in sheets:
            try:
                raw_preview = load_sheet_merge_aware(path, s)
                _st, _ed, _hdrs = detect_header_band_and_build(raw_preview)
                cand_name_headers.append((s, _hdrs))
            except Exception:
                cand_name_headers.append((s, []))

        # 2) preselect 계산: 기존 지식> (첫파일 로직+헤더 추천)
        preselect = []
        current = self.file_sheets.get(path, [])
        if current:
            preselect = current[:]
        elif getattr(self, "pref_sheet_names", None):
            base = [(bn, self.pref_headers.get(bn, [])) for bn in self.pref_sheet_names]
            preselect = auto_match_with_headers(base, cand_name_headers, threshold=0.55)

        auto_allowed = not force_dialog
        has_preview = self._has_preview()

        # 3) 시트가 1개뿐이면 (미리보기 여부와 무관하게) 팝업 없이 즉시 적용
        if auto_allowed and len(sheets) == 1:
            chosen = sheets[:]
            self._record_pref_from(path, chosen)
            self._finalize_file_selection(path, chosen)
            self._parse_and_preview()
            return

        # 3-1) 추천 결과가 단일 시트라면 팝업 없이 자동 적용
        if auto_allowed and has_preview and len(preselect) == 1:
            chosen = preselect[:]
            self._finalize_file_selection(path, chosen)
            self._parse_and_preview()
            return

        # 4) 후보가 0개 또는 2개 이상)팝업 유도: 여러 개면 여러 개를 체크상태로
        dlg = SheetChooser(self, path, sheets, preselect)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        chosen = dlg.selected_sheets()
        if not chosen:
            return

        # 첫파일의 첫 선택이면 기준 시트/헤더 저장
        self._record_pref_from(path, chosen)

        # 파일 목록/시트 매핑 반영
        self._finalize_file_selection(path, chosen)

        # 즉시 미리보기 반영
        self._parse_and_preview()
    # ---------- Core parse & combine ----------
    def _clear(self):
        self.file_paths.clear()
        self.file_sheets.clear()
        self.list.clear()
        self.parsed.clear()
        self.combined = None
        self.saved_header_signature = None
        self.saved_header_band = None
        self.sheet_header_cache = {}
        self.pref_sheet_names = []
        self.pref_headers = {}
        self.table.setModel(None)

    def _parse_and_preview(self):
        if not self.file_paths:
            self.table.setModel(None)
            self.parsed.clear()
            self.combined = None
            self.saved_header_signature = None
            self.saved_header_band = None
            self.sheet_header_cache = {}
            return
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.progress.setMaximum(len(self.file_paths))

        parsed: List[ParsedSheet] = []
        for i, p in enumerate(self.file_paths, 1):
            try:
                sheets = self.file_sheets.get(p) or list_sheet_names(p)[:1]
                for sh in sheets:
                    raw = load_sheet_merge_aware(p, sh)
                    start, end, headers = detect_header_band_and_build(raw)
                    start, end, headers = self._resolve_header_band(raw, p, sh, start, end)
                    pd = _pd()
                    data = raw.iloc[end+1:, :].copy()
                    data.columns = headers[: data.shape[1]]
                    if len(headers) > data.shape[1]:
                        for j in range(data.shape[1], len(headers)):
                            data[headers[j]] = pd.NA
                    data = data.dropna(how='all')
                    for c in data.columns:
                        data[c] = data[c].apply(lambda x: str(x).strip() if (x is not None and str(x).strip() != "nan") else x)
                    parsed.append(ParsedSheet(file_path=p, sheet_name=sh, columns=headers, data=data))
            except Exception as e:
                traceback.print_exc()
                QMessageBox.critical(self, "읽기 오류", f"{os.path.basename(p)} 처리 중 오류\n{e}")
            finally:
                self.progress.setValue(i)
        self.parsed = parsed

        if not parsed:
            QMessageBox.warning(self, "알림", "유효한 시트를 찾지 못했습니다.")
            self.progress.setVisible(False)
            return

        pd = _pd()
        all_cols: List[str] = []
        for ps in parsed:
            for c in ps.data.columns:
                if c not in all_cols:
                    all_cols.append(c)

        frames = []
        for ps in parsed:
            df = ps.data.copy()
            for c in all_cols:
                if c not in df.columns:
                    df[c] = pd.NA
            df = df[all_cols]
            df[META_MAP["file"]] = os.path.basename(ps.file_path)
            df[META_MAP["sheet"]] = str(ps.sheet_name)
            frames.append(df)

        combined = _pd().concat(frames, ignore_index=True)
        combined = combined.dropna(how='all', subset=[c for c in all_cols])
        combined[META_MAP["violations"]] = compute_violations(combined)

        self.combined = combined
        self.progress.setVisible(False)
        self._refresh_preview()

    def _reselect_header(self):
        selected = self.list.selectedItems()
        if not selected:
            # 선택이 없고 파일이 1개뿐이면 자동 선택
            if self.list.count() == 1:
                self.list.setCurrentRow(0)
            else:
                QMessageBox.information(self, "알림", "헤더를 다시 선택할 파일을 목록에서 선택하세요.")
                return
        # 시그니처 초기화 후 재파싱(보정 UI를 다시 띄움)
        self.saved_header_signature = None
        self.saved_header_band = None
        self.sheet_header_cache = {}
        self._parse_and_preview()

    def _filtered_for_preview(self):
        if self.combined is None:
            return None
        df = self.combined.copy()
        if not self.chk_meta.isChecked():
            drop_cols = [c for c in META_COLS_KR if c in df.columns]
            if drop_cols:
                df = df.drop(columns=drop_cols)
        return df.head(500).copy()

    def _refresh_preview(self):
        if self.combined is None:
            return
        preview = self._filtered_for_preview()
        self.table.setModel(PandasModel(preview))

    def _save_output(self):
        if self.combined is None or len(self.combined) == 0:
            QMessageBox.warning(self, "알림", "먼저 파일을 드래그하여 미리보기를 확인하세요.")
            return
        default_name = "취합_결과.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "저장 위치 선택", default_name, "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            pd = _pd()
            df_out = self.combined.copy()
            if not self.chk_meta.isChecked():
                drop_cols = [c for c in META_COLS_KR if c in df_out.columns]
                if drop_cols:
                    df_out = df_out.drop(columns=drop_cols)

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "합본"

            split_cols = [str(c).split(" | ") for c in df_out.columns]
            depth = max(len(parts) for parts in split_cols) if split_cols else 1
            depth = min(max(depth, 1), 5)

            norm_cols = []
            for parts in split_cols:
                parts = parts[-depth:]
                if len(parts) < depth:
                    parts = ([""] * (depth - len(parts))) + parts
                norm_cols.append(parts)

            for r in range(depth):
                for c, parts in enumerate(norm_cols, start=1):
                    val = parts[r] if parts[r] != "" else None
                    ws.cell(row=r+1, column=c, value=val)

            for r in range(depth):
                cur, start = None, 1
                for c in range(1, len(norm_cols)+1):
                    v = ws.cell(row=r+1, column=c).value or ""
                    if cur is None:
                        cur, start = v, c
                    elif v != cur:
                        if cur != "" and (c-1) > start:
                            ws.merge_cells(start_row=r+1, start_column=start, end_row=r+1, end_column=c-1)
                        cur, start = v, c
                if cur is not None and cur != "" and len(norm_cols) >= start+1:
                    ws.merge_cells(start_row=r+1, start_column=start, end_row=r+1, end_column=len(norm_cols))

            for c in range(1, len(norm_cols)+1):
                for r in range(1, depth):
                    top = ws.cell(row=r, column=c).value
                    bot = ws.cell(row=r+1, column=c).value
                    if (top is not None and str(top) != "") and (bot is None or str(bot) == ""):
                        ws.merge_cells(start_row=r, start_column=c, end_row=r+1, end_column=c)

            for i, row in enumerate(df_out.itertuples(index=False), start=depth+1):
                for j, val in enumerate(row, start=1):
                    try:
                        import pandas as pd
                        ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
                    except Exception:
                        ws.cell(row=i, column=j, value=val)

            wb.save(path)

            log_path = None
            if self.chk_log.isChecked():
                log_path = os.path.join(os.path.dirname(path), "취합_로그.csv")
                _pd().DataFrame([
                    {"file": os.path.basename(p.file_path), "sheet": p.sheet_name}
                    for p in self.parsed
                ]).to_csv(log_path, index=False, encoding="utf-8-sig")

            msg = f"출력: {path}"
            if log_path:
                msg += f"\n로그: {log_path}"
            QMessageBox.information(self, "저장됨", msg)
        except Exception as err:
            traceback.print_exc()
            QMessageBox.critical(self, "저장 오류", str(err))
            
class HeaderAdjustDialog(QDialog):
    def __init__(self, parent, df_raw, guess_start: int, guess_end: int):
        super().__init__(parent)
        self.setWindowTitle("헤더 보정")
        self.resize(900, 500)
        layout = QVBoxLayout(self)

        guide = QLabel("원하는 영역에서 헤더로 쓸 행을 드래그해 선택하고, 위/아래 버튼으로 크기를 조정하세요. 선택된 행만 헤더로 적용됩니다.")
        guide.setStyleSheet("color:#444;padding:4px 0;")
        layout.addWidget(guide)

        top_rows = max(10, min(60, len(df_raw)))
        df_prev = df_raw.head(top_rows).copy()
        try:
            df_prev = df_prev.fillna("")
        except Exception:
            pass

        self.model = PandasModel(df_prev)
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableView.SelectionMode.ExtendedSelection)
        self.table.setStyleSheet(
            "QTableView::item:selected{background-color:#ffe6b3;color:#000;border:1px solid #ff8c00;}"
        )
        layout.addWidget(self.table, 1)

        ctl = QHBoxLayout()
        ctl.addWidget(QLabel("헤더 시작 행"))
        self.spin_start = QSpinBox()
        self.spin_start.setMinimum(1)
        self.spin_start.setMaximum(top_rows)
        self.spin_start.setValue(min(max(1, guess_start+1), top_rows))
        ctl.addWidget(self.spin_start)

        ctl.addWidget(QLabel("헤더 끝 행"))
        self.spin_end = QSpinBox()
        self.spin_end.setMinimum(1)
        self.spin_end.setMaximum(top_rows)
        self.spin_end.setValue(min(max(1, guess_end+1), top_rows))
        ctl.addWidget(self.spin_end)

        ctl.addStretch(1)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        ctl.addWidget(btns)
        layout.addLayout(ctl)

        self._sync_guard = False
        sel_model = self.table.selectionModel()
        sel_model.selectionChanged.connect(lambda *_: self._selection_changed())
        self.spin_start.valueChanged.connect(lambda _=None: self._sync_selection("start"))
        self.spin_end.valueChanged.connect(lambda _=None: self._sync_selection("end"))
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        self._sync_selection()

    def _sync_selection(self, source=None):
        s_val = self.spin_start.value()
        e_val = self.spin_end.value()
        if source == "start" and s_val > e_val:
            e_val = s_val
            self.spin_end.blockSignals(True)
            self.spin_end.setValue(e_val)
            self.spin_end.blockSignals(False)
        elif source == "end" and e_val < s_val:
            s_val = e_val
            self.spin_start.blockSignals(True)
            self.spin_start.setValue(s_val)
            self.spin_start.blockSignals(False)
        s = s_val - 1
        e = e_val - 1
        sel = self.table.selectionModel()
        if sel:
            self._sync_guard = True
            sel.clearSelection()
            model = self.table.model()
            if model and model.rowCount() > 0:
                max_row = model.rowCount() - 1
                max_col = max(0, model.columnCount() - 1)
                s_clamped = max(0, min(s, max_row))
                e_clamped = max(0, min(e, max_row))
                top_left = model.index(s_clamped, 0)
                bottom_right = model.index(e_clamped, max_col)
                sel_range = QItemSelection(top_left, bottom_right)
                sel.select(sel_range, QItemSelectionModel.SelectionFlag.Select | QItemSelectionModel.SelectionFlag.Rows)
            self._sync_guard = False
        try:
            self.table.scrollTo(self.table.model().index(s, 0))
        except Exception:
            pass
        self.table.viewport().update()

    def _selection_changed(self):
        if getattr(self, "_sync_guard", False):
            return
        sel = self.table.selectionModel()
        if not sel:
            return
        rows = [i.row() for i in sel.selectedRows()]
        if not rows:
            return
        s, e = min(rows), max(rows)
        # block signals to avoid recursion
        self.spin_start.blockSignals(True)
        self.spin_end.blockSignals(True)
        self.spin_start.setValue(s+1)
        self.spin_end.setValue(e+1)
        self.spin_start.blockSignals(False)
        self.spin_end.blockSignals(False)

    def selected_band(self) -> tuple[int, int]:
        s = min(self.spin_start.value(), self.spin_end.value()) - 1
        e = max(self.spin_start.value(), self.spin_end.value()) - 1
        return s, e


class SheetChooser(QDialog):
    """시트 목록을 체크박스로 보여주는 단순 다이얼로그.
       preselect에 포함된 시트는 미리 체크됨.
    """
    def __init__(self, parent, file_path: str, sheets: list[str], preselect: list[str] | None = None):
        super().__init__(parent)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.setWindowTitle(f"시트 선택 ― {os.path.basename(file_path)}")
        self.resize(420, 520)
        self._sheets = sheets
        self._pre = set(preselect or [])

        layout = QVBoxLayout(self)
        self.info = QLabel("취합할 시트를 선택하세요. (복수 선택 가능)")
        self.info.setStyleSheet("color:#333;")
        layout.addWidget(self.info)

        self.listw = QListWidget(self)
        for s in sheets:
            it = QListWidgetItem(s)
            flags = it.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
            it.setFlags(flags)
            if s in self._pre:
                it.setCheckState(Qt.CheckState.Checked)
            else:
                it.setCheckState(Qt.CheckState.Unchecked)
            self.listw.addItem(it)
        layout.addWidget(self.listw, 1)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            parent=self
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        QTimer.singleShot(0, self._ensure_attention)

    def selected_sheets(self) -> list[str]:
        out: list[str] = []
        for i in range(self.listw.count()):
            it = self.listw.item(i)
            if it.checkState() == Qt.CheckState.Checked:
                out.append(it.text())
        return out

    def _ensure_attention(self):
        try:
            self.raise_()
            self.activateWindow()
            QApplication.alert(self, 0)
        except Exception:
            pass

class SingleSheetDialog(QDialog):
    def __init__(self, parent, file_path: str, sheets: list[str]):
        super().__init__(parent)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, True)
        self.setWindowTitle(f"헤더 확인/조정 – {os.path.basename(file_path)}")
        self.resize(360, 400)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("헤더를 조정할 시트를 선택하세요."))
        self.listw = QListWidget(self)
        self.listw.addItems(sheets)
        self.listw.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        layout.addWidget(self.listw, 1)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(btns)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        if sheets:
            self.listw.setCurrentRow(0)

    def selected_sheet(self) -> Optional[str]:
        it = self.listw.currentItem()
        return it.text() if it else None

def main():
    os.environ.setdefault("QT_ENABLE_HIGHDPI_SCALING", "1")
    os.environ.setdefault("QT_SCALE_FACTOR", "1")
    app = QApplication(sys.argv)
    w = ExcelAggregator()
    w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
